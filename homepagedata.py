import openpyxl

class homepgdata:
    test_homepgdata=[{"madhavi","madd@","madhu"},{"ganesh","gan@","gane"}]
    @staticmethod
    def gettestdata():
        book = openpyxl.load_workbook("C:\\Users\\ganesh ramba\\Downloads\\exceldemo.xlsx")
        sheet = book.active
        dict={}
        for i in range(2, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                dict[(sheet.cell(row=1, column=j).value)] = sheet.cell(row=i, column=j).value
            return [dict]

